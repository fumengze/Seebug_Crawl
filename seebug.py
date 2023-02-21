from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.by import By
import time,os
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as ws


# chrome配置函数
def start_chrom():
	print('[+] 正在后台打开谷歌浏览器...')
	chrome_option = Options()
	# chrome_option.add_argument('blink-settings=imagesEnabled=false')#关闭图片显示
	chrome_option.add_argument("--disable-blink-features=AutomationControlled")#绕过反爬参数
	chrome_option.add_argument('--ignore-certificate-errors')#关闭错误拦截，您的连接不是私密连接
	# chrome_option.add_argument('--headless')#开启无头浏览器
	#chrome_option.add_argument('--proxy-server=127.0.0.1:10809')#添加代理，只支持http和https
	chrome_option.add_experimental_option('excludeSwitches', ['enable-logging'])  # 关闭控制台日志
	browser = webdriver.Chrome(executable_path="D:\\security\\pychrom2021\\Projects\\项目\\测试目录_demo\\brute_force\\chrome_bruteV0.3\\chromedriver.exe", chrome_options=chrome_option)
	browser.set_page_load_timeout(500)
	return browser

def creat_xlsx():
	filename = 'out\\%s.xlsx' % time.strftime("%Y-%m-%d-%H-%M", time.localtime(time.time()))
	if os.path.exists(filename) == False:
		s = 0
		wb = ws.Workbook()
		ws1 = wb.active
		if os.path.exists('out\\') == False:
			os.mkdir('out')
		word=['漏洞名称','seebug地址','漏洞POC地址','CVE编号','CNNVD编号','CNVD编号']
		for i in word:
			s = s + 1
			ws1.cell(row =1,column = s,value = i)
		wb.save(filename)
		print("[*]创建文件成功 %s"%filename)
		return filename
	else:
		print("[*]文件已存在 文件为:%s"%filename)
		return filename

def write_xlsx(poc_name,seebug_url,poc_url,CVE,CNNVD,CNVD,wb):
	# print(define.GREEN+"[*]内容正在写入 :%s"%res)
	sheet1 = wb['Sheet']
	num = sheet1.max_row
	print(num)
	sheet1.cell(row = num+1,column = 1,value = poc_name)
	sheet1.cell(row = num+1,column = 2,value = seebug_url)
	sheet1.cell(row = num+1,column = 3,value = poc_url)
	sheet1.cell(row=num + 1, column=4, value=CVE)
	sheet1.cell(row=num + 1, column=5, value=CNNVD)
	sheet1.cell(row=num + 1, column=6, value=CNVD)

def login(browser):
	browser.get("https://sso.telnet404.com/cas/login?service=http%3A%2F%2Fwww.seebug.org%2Faccounts%2Flogin%2F%3Fnext%3D%252Fvuldb%252Fvulnerabilities%253Fhas_poc%253Dtrue%2526page%253D1")
	time.sleep(20)
	try:
		print(browser.get_cookies())
		for cookie in browser.get_cookies():
			print(cookie)
			browser.add_cookie(cookie)
		time.sleep(1)
		return browser
	except Exception as e:  # 出现其他未知异常时抛出异常
		print(e)



def requests_content(browser,filename):
	wait = WebDriverWait(browser, 10)
	try:
		vul_name = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//a[@class="vul-title"]')))
		print("[+]数量："+str(len(vul_name)))
	except Exception as e:
		print("[-]获取漏洞名称时出现异常！！！" + str(e))
	wb = ws.load_workbook(filename)
	for name in vul_name:
		time.sleep(10)
		poc_name = name.text
		seebug_url = name.get_attribute('href')
		print("[+]漏洞名称："+poc_name)
		print("[+]seebug地址："+seebug_url)
		browser.execute_script("window.open('')")#打开新的tab标签页
		browser.switch_to.window(browser.window_handles[1])#切换标签页
		time.sleep(1)
		browser.get(seebug_url)#请求漏洞URL
		try:
			vul_url = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[2]/div[1]/section[1]/div/div/a')))
			poc_url = vul_url.text
			print("[+]漏洞POC地址："+poc_url)
		except Exception as e:
			print("[-]没有获取到漏洞来源地址！！！"+str(e))
			poc_url = "null"
		try:
			CVE = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[2]/div[1]/div/section/div/div[3]/dl[1]/dd/a'))).text
			CNNVD = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[2]/div[1]/div/section/div/div[3]/dl[2]/dd/a'))).text
			CNVD = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[2]/div[1]/div/section/div/div[3]/dl[3]/dd/a'))).text
			print("[+]CVE编号："+CVE+'	CNNVD编号：'+CNNVD+'	CNVD编号：'+CNVD)
		except Exception as e:
			print("[-]没有找到漏洞编号！！！")
			print(e)
			CVE = "null"
			CNNVD = "null"
			CNVD = "null"
		browser.close()
		# browser.switch_to.window(browser.window_handles[1])
		# browser.close()
		browser.switch_to.window(browser.window_handles[0])
		time.sleep(1)
		try:
			write_xlsx(poc_name,seebug_url,poc_url,CVE,CNNVD,CNVD,wb)
			time.sleep(1)
		except Exception as e:
			print(e)
			time.sleep(1)
	wb.save(filename)
	wb.close()

if __name__ == "__main__":
	browser = start_chrom()
	browser = login(browser)
	filename = creat_xlsx()  # 初次执行时创建xlsx文件
	for page in range(1,2928):
		print(page)
		# target = "https://www.seebug.org/vuldb/vulnerabilities?has_poc=true&page={page}".format(page=page)
		target = "https://www.seebug.org/vuldb/vulnerabilities?page={page}".format(page=page)
		browser.get(target)
		requests_content(browser,filename)
	browser.close()